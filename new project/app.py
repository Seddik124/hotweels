import os
import pandas as pd
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

app = Flask(__name__, template_folder='templates')

@app.route('/')
def index():
    return render_template('index.html')  # Assure-toi que ce nom est exact

@app.route('/get_sheets', methods=['POST'])
def get_sheets():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file provided'}), 400

    filename = secure_filename(file.filename)
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(file_path)

    try:
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        os.remove(file_path)  # Clean up the temporary file
        return jsonify({'sheets': sheet_names})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/process', methods=['POST'])
def process_file():
    file = request.files.get('file')
    action = request.form.get('action')
    sheet_name = request.form.get('sheet_name')
    if not file or not action or not sheet_name or sheet_name == 'null':
        return jsonify({'error': 'Fichier, action ou nom de feuille manquant'}), 400

    filename = secure_filename(file.filename)
    upload_path = os.path.join(UPLOAD_FOLDER, filename)
    result_path = os.path.join(RESULT_FOLDER, f"{os.path.splitext(filename)[0]}_corrigé.xlsx")
    file.save(upload_path)

    try:
        if action == 'detect_errors':
            result, error_list = detect_errors(upload_path, result_path, sheet_name)
            return jsonify({
                'results': {'errors': error_list},
                'file': os.path.basename(result_path)
            })
        elif action == 'detect_duplicates':
            result, doublons = detect_duplicates(upload_path, result_path, sheet_name)
            return jsonify({
                'results': {'doublons': doublons},
                'file': os.path.basename(result_path)
            })
        else:
            return jsonify({'error': 'Action inconnue'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/download', methods=['POST'])
def download_file():
    data = request.get_json()
    file_name = data.get('file')
    file_path = os.path.join(RESULT_FOLDER, file_name)

    if not os.path.exists(file_path):
        return jsonify({'error': 'Fichier non trouvé'}), 404

    return send_file(file_path, as_attachment=True)


# ----------------------
# LOGIQUE : DÉTECTION D'ERREURS
# ----------------------
def detect_errors(input_path, output_path, sheet_name):
    header_row = 1
    data_start_row = 2
    df = pd.read_excel(input_path, sheet_name=sheet_name, header=header_row)
    df.columns = [str(col).strip() for col in df.columns]

    cols = {
        '2G': {
            'freq': "fréquences d'émission",
            'tilt': "Tits mécanques et électriques de chaque antenne",
            'pire': "Puissance isotrope rayonnée équivalente (PIRE) dans chaque secteur",
            'ant': "Nombre d'antennes",
            'azim': "azimut du rayonnement maximum dans chaque secteur"
        },
        '3G': {
            'tilt': "Tits mécanques et électriques de chaque antenne.1",
            'pire': "Puissance isotrope rayonnée équivalente (PIRE) dans chaque secteur.1",
            'ant': "Nombre d'antennes MIMO",
            'azim': "Azimut du rayonnement maximum dans chaque secteur"
        },
        '4G': {
            'tilt': "Tits mécanques et électriques de chaque antenne.2",
            'pire': "Puissance isotrope rayonnée équivalente (PIRE) dans chaque secteur.2",
            'ant': "Nombre d'antennes MIMO.1",
            'azim': "Azimut du rayonnement maximum dans chaque secteur.1"
        }
    }

    def parse_values(value):
        if pd.isna(value) or str(value).strip() in ['', 'nan']:
            return None
        try:
            parts = str(value).replace(',', '.').split('/')
            return [float(x.strip()) for x in parts if x.strip()]
        except ValueError:
            return None

    error_lines = []

    for idx, row in df.iterrows():
        errors = []
        freq_2g = parse_values(row[cols['2G']['freq']])
        if not freq_2g:
            continue

        ref_count = len(freq_2g)
        ref_azim = parse_values(row[cols['2G']['azim']])

        for gen in ['2G', '3G', '4G']:
            for field in ['tilt', 'pire', 'ant']:
                val = parse_values(row[cols[gen][field]])
                if val and len(val) != ref_count:
                    errors.append({
                        "Ligne": idx + data_start_row + 1,
                        "Colonne": cols[gen][field],
                        "Valeur": row[cols[gen][field]],
                        "Problème": f"{gen} - {field}: {len(val)} ≠ {ref_count}"
                    })
            azim = parse_values(row[cols[gen]['azim']])
            if azim and ref_azim and azim != ref_azim:
                errors.append({
                    "Ligne": idx + data_start_row + 1,
                    "Colonne": cols[gen]['azim'],
                    "Valeur": row[cols[gen]['azim']],
                    "Problème": f"{gen} - azimut ≠ 2G"
                })

        if errors:
            error_lines.extend(errors)

    # Coloration si erreur
    if error_lines:
        wb = load_workbook(input_path)
        ws = wb[sheet_name]
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        for err in error_lines:
            row_num = err["Ligne"]
            for row in ws.iter_rows(min_row=row_num, max_row=row_num):
                for cell in row:
                    cell.fill = red_fill
        wb.save(output_path)
    else:
        output_path = input_path  # Si aucun problème

    return output_path, error_lines


# ----------------------
# LOGIQUE : DÉTECTION DE DOUBLONS
# ----------------------
def detect_duplicates(input_path, output_path, sheet_name):
    df = pd.read_excel(input_path, sheet_name=sheet_name, header=2)
    df = df.rename(columns={
        df.columns[0]: 'Identifiant',
        df.columns[4]: 'Longitude',
        df.columns[5]: 'Latitude'
    })

    df['Latitude'] = pd.to_numeric(df['Latitude'], errors='coerce')
    df['Longitude'] = pd.to_numeric(df['Longitude'], errors='coerce')
    df_coords = df.dropna(subset=['Latitude', 'Longitude'])

    groupes = df_coords.groupby(['Latitude', 'Longitude'])['Identifiant'].agg(['count', 'nunique'])
    doublons = groupes[groupes['nunique'] > 1]
    lignes_doublons = []

    for (lat, lon), _ in doublons.iterrows():
        mask = (df_coords['Latitude'] == lat) & (df_coords['Longitude'] == lon)
        lignes_doublons.extend(df_coords[mask].index.tolist())

    results = []
    for idx in lignes_doublons:
        row = df.iloc[idx]
        results.append({
            "Ligne": idx + 4,
            "Identifiant": row['Identifiant'],
            "Latitude": row['Latitude'],
            "Longitude": row['Longitude']
        })

    # Colorier les lignes trouvées
    wb = load_workbook(input_path)
    ws = wb[sheet_name]
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for idx in lignes_doublons:
        row_num = idx + 4
        for row in ws.iter_rows(min_row=row_num, max_row=row_num):
            for cell in row:
                cell.fill = green_fill

    wb.save(output_path)
    return output_path, results


if __name__ == '__main__':
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)

